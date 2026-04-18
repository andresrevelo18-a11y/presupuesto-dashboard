#!/usr/bin/env python3
"""Exporta data/presupuesto.xlsx a data/presupuesto.json y data/presupuesto.js."""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "data" / "presupuesto.xlsx"
JSON_OUT = ROOT / "data" / "presupuesto.json"
JS_OUT = ROOT / "data" / "presupuesto.js"


def number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).replace("$", "").replace(",", "").strip()
    return round(float(text), 2) if text else 0.0


def read_table(ws_name: str) -> list[dict[str, Any]]:
    ws = wb[ws_name]
    headers = [cell.value for cell in ws[1]]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in row):
            continue
        rows.append({str(headers[i]): row[i] for i in range(len(headers))})
    return rows


if not WORKBOOK.exists():
    raise SystemExit(f"No existe {WORKBOOK}")

wb = load_workbook(WORKBOOK, data_only=True)

summary = {row["key"]: row["value"] for row in read_table("Resumen")}
months = read_table("Meses")
flujo = read_table("Flujo")
mayo = read_table("Mayo")
semanal = read_table("Semanal")
deudas = read_table("Deudas")
activos = read_table("Activos")

data = {
    "generatedAt": datetime.now(timezone.utc).isoformat(),
    "summary": summary,
    "months": months,
    "charts": {
        "flujo": {
            "labels": [row["mes"] for row in flujo],
            "ingresos": [number(row["ingresos"]) for row in flujo],
            "gastos": [number(row["gastos"]) for row in flujo],
            "saldo": [number(row["saldo"]) for row in flujo],
        },
        "mayo": {
            "labels": [row["categoria"] for row in mayo],
            "values": [number(row["monto"]) for row in mayo],
        },
        "semanal": {
            "labels": [row["semana"] for row in semanal],
            "values": [number(row["bolsa"]) for row in semanal],
        },
        "deudas": {
            "labels": [row["deuda"] for row in deudas],
            "values": [number(row["saldo"]) for row in deudas],
        },
        "activos": {
            "labels": [row["activo"] for row in activos],
            "values": [number(row["valor"]) for row in activos],
        },
    },
}

JSON_OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
JS_OUT.write_text(
    "window.PRESUPUESTO_DATA = "
    + json.dumps(data, ensure_ascii=False, indent=2)
    + ";\n",
    encoding="utf-8",
)

print(f"Exportado {JSON_OUT.relative_to(ROOT)}")
print(f"Exportado {JS_OUT.relative_to(ROOT)}")
