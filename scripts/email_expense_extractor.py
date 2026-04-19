#!/usr/bin/env python3
"""
Agente que revisa el correo de Gmail y extrae gastos a una tabla Excel.

Uso:
  python scripts/email_expense_extractor.py
  python scripts/email_expense_extractor.py --max-emails 50 --days 30
  python scripts/email_expense_extractor.py --output mi_salida.xlsx

Requisitos:
  pip install anthropic google-auth-oauthlib google-auth-httplib2 google-api-python-client openpyxl

Primera vez: el script abrirá el navegador para autenticar con Gmail.
Las credenciales quedan guardadas en token.json para futuros usos.
"""

from __future__ import annotations

import argparse
import base64
import json
import os
import re
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import anthropic
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ─────────────────────────────── Google imports ──────────────────────────────
try:
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build
except ImportError:
    sys.exit(
        "Faltan dependencias. Instala con:\n"
        "  pip install google-auth-oauthlib google-auth-httplib2 google-api-python-client"
    )

# ─────────────────────────────── Constantes ──────────────────────────────────
SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]
ROOT = Path(__file__).resolve().parents[1]
TOKEN_PATH = ROOT / "token.json"
CREDENTIALS_PATH = ROOT / "credentials.json"

# Columnas del Excel de salida
COLUMNS = [
    "Fecha",
    "Monto",
    "Moneda",
    "Comercio / Proveedor",
    "Descripción",
    "Categoría",
    "Asunto del correo",
    "Fecha del correo",
]

# Categorías disponibles para que Claude clasifique
CATEGORIES = [
    "Alimentación",
    "Transporte",
    "Salud",
    "Entretenimiento",
    "Servicios (agua/luz/internet/teléfono)",
    "Ropa y calzado",
    "Educación",
    "Viajes y hospedaje",
    "Suscripciones",
    "Compras en línea",
    "Restaurantes y cafeterías",
    "Supermercado",
    "Seguros",
    "Inversiones / Ahorro",
    "Otro",
]

EXPENSE_QUERY = (
    "subject:(recibo OR factura OR comprobante OR receipt OR invoice OR "
    "order OR pedido OR pago OR cobro OR cargo OR purchase OR payment OR "
    "\"tu compra\" OR \"tu pedido\" OR \"confirmación de pago\")"
)


# ─────────────────────────────── Gmail auth ──────────────────────────────────

def get_gmail_service():
    """Autentica con Gmail y devuelve el servicio de la API."""
    creds: Credentials | None = None

    if TOKEN_PATH.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_PATH), SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not CREDENTIALS_PATH.exists():
                sys.exit(
                    f"No se encontró {CREDENTIALS_PATH}.\n"
                    "Descarga 'credentials.json' desde Google Cloud Console:\n"
                    "  https://console.cloud.google.com/apis/credentials\n"
                    "Crea un proyecto, habilita Gmail API y descarga las credenciales OAuth 2.0."
                )
            flow = InstalledAppFlow.from_client_secrets_file(
                str(CREDENTIALS_PATH), SCOPES
            )
            creds = flow.run_local_server(port=0)
        TOKEN_PATH.write_text(creds.to_json())
        print(f"✓ Credenciales guardadas en {TOKEN_PATH.relative_to(ROOT)}")

    return build("gmail", "v1", credentials=creds)


# ─────────────────────────────── Email fetch ─────────────────────────────────

def fetch_emails(service, max_results: int = 100, days_back: int = 60) -> list[dict]:
    """Obtiene emails relevantes de los últimos `days_back` días."""
    after_date = (datetime.now() - timedelta(days=days_back)).strftime("%Y/%m/%d")
    query = f"{EXPENSE_QUERY} after:{after_date}"

    print(f"🔍 Buscando correos de los últimos {days_back} días...")
    result = (
        service.users()
        .messages()
        .list(userId="me", q=query, maxResults=max_results)
        .execute()
    )
    messages = result.get("messages", [])
    print(f"   Encontrados: {len(messages)} correos candidatos")

    emails: list[dict] = []
    for i, msg_ref in enumerate(messages, 1):
        msg = (
            service.users()
            .messages()
            .get(userId="me", id=msg_ref["id"], format="full")
            .execute()
        )
        email_data = _parse_email(msg)
        if email_data:
            emails.append(email_data)
        if i % 10 == 0:
            print(f"   Procesando... {i}/{len(messages)}")

    return emails


def _parse_email(msg: dict) -> dict | None:
    """Extrae asunto, fecha y texto plano de un mensaje de Gmail."""
    headers = {h["name"]: h["value"] for h in msg["payload"].get("headers", [])}
    subject = headers.get("Subject", "(sin asunto)")
    date_str = headers.get("Date", "")

    body = _extract_body(msg["payload"])
    if not body.strip():
        return None

    # Limitar a 4000 caracteres para no sobrecargar el contexto de Claude
    body_snippet = body[:4000]

    return {
        "id": msg["id"],
        "subject": subject,
        "date": date_str,
        "body": body_snippet,
    }


def _extract_body(payload: dict) -> str:
    """Extrae texto plano del payload de un mensaje (recursivo en partes)."""
    mime_type = payload.get("mimeType", "")
    body_data = payload.get("body", {}).get("data", "")

    if mime_type == "text/plain" and body_data:
        return base64.urlsafe_b64decode(body_data).decode("utf-8", errors="replace")

    if mime_type in ("multipart/alternative", "multipart/mixed", "multipart/related"):
        for part in payload.get("parts", []):
            text = _extract_body(part)
            if text:
                return text

    # Fallback: HTML → extraer texto sin etiquetas
    if mime_type == "text/html" and body_data:
        html = base64.urlsafe_b64decode(body_data).decode("utf-8", errors="replace")
        return re.sub(r"<[^>]+>", " ", html)

    return ""


# ─────────────────────────────── Claude extraction ───────────────────────────

SYSTEM_PROMPT = """Eres un asistente experto en finanzas personales.
Analizas correos electrónicos para extraer información de gastos o compras.

Reglas:
- Solo extrae gastos reales (compras, pagos, facturas, recibos). Ignora publicidad pura o newsletters sin transacción.
- Si el correo contiene múltiples items, extrae cada uno como gasto separado.
- Usa la moneda que aparezca en el correo (USD, EUR, COP, MXN, PEN, etc.). Si no hay símbolo, usa USD.
- Para la fecha, usa el formato YYYY-MM-DD. Si no hay fecha de transacción, usa la fecha del correo.
- El monto debe ser numérico (sin símbolos de moneda).
- Clasifica en una de estas categorías exactas: """ + ", ".join(CATEGORIES) + """

Si el correo NO contiene ningún gasto real, responde exactamente: {"gastos": []}

Responde SIEMPRE con JSON válido en este formato:
{
  "gastos": [
    {
      "fecha": "YYYY-MM-DD",
      "monto": 29.99,
      "moneda": "USD",
      "comercio": "Nombre del comercio o proveedor",
      "descripcion": "Descripción breve del gasto",
      "categoria": "Categoría exacta de la lista"
    }
  ]
}"""


def extract_expenses_from_email(
    client: anthropic.Anthropic, email: dict
) -> list[dict]:
    """Usa Claude para extraer gastos de un correo."""
    user_message = f"""Correo:
Asunto: {email["subject"]}
Fecha: {email["date"]}

Contenido:
{email["body"]}"""

    try:
        response = client.messages.create(
            model="claude-opus-4-7",
            max_tokens=1024,
            thinking={"type": "adaptive"},
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": user_message}],
        )

        # Obtener el bloque de texto de la respuesta
        text = next(
            (b.text for b in response.content if b.type == "text"), ""
        )
        if not text.strip():
            return []

        # Parsear JSON (puede venir entre backticks de markdown)
        json_match = re.search(r"\{.*\}", text, re.DOTALL)
        if not json_match:
            return []

        data = json.loads(json_match.group())
        return data.get("gastos", [])

    except (json.JSONDecodeError, anthropic.APIError) as e:
        print(f"   ⚠ Error procesando '{email['subject'][:50]}': {e}")
        return []


# ─────────────────────────────── Excel output ────────────────────────────────

def write_excel(rows: list[dict], output_path: Path) -> None:
    """Escribe los gastos extraídos en un archivo Excel con formato."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos del correo"

    # Encabezados
    header_fill = PatternFill("solid", fgColor="2E4057")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_num, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Datos
    alt_fill = PatternFill("solid", fgColor="F0F4F8")
    for row_num, gasto in enumerate(rows, 2):
        fill = alt_fill if row_num % 2 == 0 else None
        values = [
            gasto.get("fecha", ""),
            gasto.get("monto", ""),
            gasto.get("moneda", ""),
            gasto.get("comercio", ""),
            gasto.get("descripcion", ""),
            gasto.get("categoria", ""),
            gasto.get("_asunto", ""),
            gasto.get("_fecha_correo", ""),
        ]
        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=col_num in (5, 7))
            if fill:
                cell.fill = fill

    # Formato de columna Monto como número
    monto_col = get_column_letter(2)
    for row_num in range(2, len(rows) + 2):
        ws[f"{monto_col}{row_num}"].number_format = "#,##0.00"

    # Anchos de columna
    col_widths = [12, 12, 10, 28, 40, 28, 45, 15]
    for col_num, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Congelar fila de encabezado
    ws.freeze_panes = "A2"

    # Hoja de resumen por categoría
    ws_resumen = wb.create_sheet("Resumen por categoría")
    category_totals: dict[str, float] = {}
    for gasto in rows:
        cat = gasto.get("categoria", "Otro")
        try:
            monto = float(gasto.get("monto", 0))
        except (TypeError, ValueError):
            monto = 0.0
        category_totals[cat] = category_totals.get(cat, 0.0) + monto

    ws_resumen.cell(row=1, column=1, value="Categoría").font = Font(bold=True)
    ws_resumen.cell(row=1, column=2, value="Total (moneda mixta)").font = Font(bold=True)
    for i, (cat, total) in enumerate(
        sorted(category_totals.items(), key=lambda x: -x[1]), 2
    ):
        ws_resumen.cell(row=i, column=1, value=cat)
        ws_resumen.cell(row=i, column=2, value=round(total, 2))
    ws_resumen.column_dimensions["A"].width = 35
    ws_resumen.column_dimensions["B"].width = 20

    wb.save(output_path)


# ─────────────────────────────── Main ────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extrae gastos de Gmail a una tabla Excel usando Claude."
    )
    parser.add_argument(
        "--max-emails", type=int, default=100,
        help="Máximo de correos a analizar (default: 100)"
    )
    parser.add_argument(
        "--days", type=int, default=60,
        help="Días hacia atrás para buscar correos (default: 60)"
    )
    parser.add_argument(
        "--output", type=str, default="",
        help="Nombre del archivo Excel de salida (default: gastos_correo_FECHA.xlsx)"
    )
    args = parser.parse_args()

    # Nombre del archivo de salida
    if args.output:
        output_path = ROOT / "data" / args.output
    else:
        fecha = datetime.now().strftime("%Y%m%d_%H%M")
        output_path = ROOT / "data" / f"gastos_correo_{fecha}.xlsx"

    print("=" * 60)
    print("  EXTRACTOR DE GASTOS DEL CORREO")
    print("=" * 60)

    # 1. Autenticar con Gmail
    print("\n📧 Conectando con Gmail...")
    service = get_gmail_service()
    print("   ✓ Conectado")

    # 2. Obtener correos
    emails = fetch_emails(service, max_results=args.max_emails, days_back=args.days)
    if not emails:
        print("\n⚠  No se encontraron correos de gastos en el período indicado.")
        return

    # 3. Inicializar cliente Claude
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        sys.exit(
            "Falta la variable de entorno ANTHROPIC_API_KEY.\n"
            "Configúrala con: export ANTHROPIC_API_KEY='tu-clave'"
        )
    claude_client = anthropic.Anthropic(api_key=api_key)

    # 4. Extraer gastos con Claude
    print(f"\n🤖 Analizando {len(emails)} correos con Claude...")
    all_expenses: list[dict] = []
    found_count = 0

    for i, email in enumerate(emails, 1):
        expenses = extract_expenses_from_email(claude_client, email)
        for exp in expenses:
            exp["_asunto"] = email["subject"]
            exp["_fecha_correo"] = email["date"]
            all_expenses.append(exp)
        if expenses:
            found_count += 1

        if i % 10 == 0 or i == len(emails):
            print(
                f"   {i}/{len(emails)} correos analizados — "
                f"{len(all_expenses)} gastos encontrados"
            )

    if not all_expenses:
        print("\n⚠  No se encontraron gastos en los correos analizados.")
        return

    # 5. Ordenar por fecha
    def sort_key(g: dict) -> str:
        return str(g.get("fecha", ""))

    all_expenses.sort(key=sort_key)

    # 6. Escribir Excel
    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_excel(all_expenses, output_path)

    print(f"\n{'=' * 60}")
    print(f"  ✅ COMPLETADO")
    print(f"{'=' * 60}")
    print(f"  Correos analizados : {len(emails)}")
    print(f"  Correos con gastos : {found_count}")
    print(f"  Gastos extraídos   : {len(all_expenses)}")
    print(f"  Archivo guardado   : {output_path.relative_to(ROOT)}")
    print(f"{'=' * 60}\n")


if __name__ == "__main__":
    main()
